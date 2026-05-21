from flask import Blueprint, render_template, request, send_file
from flask_login import login_required, current_user
from dao.tarea_dao import TareaDAO
from dao.campana_dao import CampanaDAO
from dao.influencer_dao import InfluencerDAO
from dao.ecommerce_dao import EcommerceDAO
from dao.usuario_dao import UsuarioDAO
from decorators import requiere_permiso
from datetime import datetime, date
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

reportes_bp = Blueprint('reportes', __name__)


def _coerce_fecha(fecha_val):
    if not fecha_val:
        return None
    if isinstance(fecha_val, (datetime, date)):
        return fecha_val
    if isinstance(fecha_val, str):
        txt = fecha_val.strip()
        for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(txt, fmt)
            except ValueError:
                continue
    return None


def crear_header_footer(canvas_obj, doc):
    """Header y footer profesional para todos los PDFs"""
    canvas_obj.saveState()

    # Header con gradiente azul
    canvas_obj.setFillColor(colors.HexColor('#1e40af'))
    canvas_obj.rect(0, A4[1] - 0.6 * inch, A4[0], 0.6 * inch, fill=True, stroke=False)

    canvas_obj.setFillColor(colors.whitesmoke)
    canvas_obj.setFont('Helvetica-Bold', 18)
    canvas_obj.drawString(40, A4[1] - 0.4 * inch, 'MYLAN SYSTEM')

    # Footer con línea y fecha
    canvas_obj.setStrokeColor(colors.HexColor('#e5e7eb'))
    canvas_obj.setLineWidth(1)
    canvas_obj.line(40, 0.6 * inch, A4[0] - 40, 0.6 * inch)

    canvas_obj.setFillColor(colors.HexColor('#6b7280'))
    canvas_obj.setFont('Helvetica', 8)
    canvas_obj.drawString(40, 0.4 * inch, f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    canvas_obj.drawRightString(A4[0] - 40, 0.4 * inch, f'Página {doc.page}')

    canvas_obj.restoreState()


@reportes_bp.route('/')
@login_required
@requiere_permiso('reportes')
def index():
    """Página principal de reportes"""
    return render_template('reportes/index.html')


# ==================== REPORTES DE TAREAS ====================

@reportes_bp.route('/tareas')
@login_required
@requiere_permiso('reportes')
def tareas():
    """Vista previa de reporte de tareas"""
    estado = request.args.get('estado')
    usuario_id = request.args.get('usuario_id')
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    # Obtener tareas filtradas
    if current_user.es_admin():
        tareas = TareaDAO.listar(estado=estado, usuario_id=int(usuario_id) if usuario_id else None)
        usuarios = UsuarioDAO.listar()
    else:
        tareas = TareaDAO.listar(estado=estado, usuario_id=current_user.id)
        usuarios = []

    # Filtrar por fechas si se especifican
    if fecha_desde:
        tareas = [t for t in tareas if t.fecha_entrega and str(t.fecha_entrega) >= fecha_desde]
    if fecha_hasta:
        tareas = [t for t in tareas if t.fecha_entrega and str(t.fecha_entrega) <= fecha_hasta]

    # Estadísticas
    total = len(tareas)
    completadas = len([t for t in tareas if t.estado == 'completada'])
    pendientes = len([t for t in tareas if t.estado == 'pendiente'])
    en_proceso = len([t for t in tareas if t.estado == 'en_proceso'])
    postergadas = len([t for t in tareas if t.estado == 'postergada'])

    return render_template('reportes/tareas.html',
                           tareas=tareas,
                           usuarios=usuarios,
                           total=total,
                           completadas=completadas,
                           pendientes=pendientes,
                           en_proceso=en_proceso,
                           postergadas=postergadas,
                           filtros={'estado': estado, 'usuario_id': usuario_id,
                                    'fecha_desde': fecha_desde, 'fecha_hasta': fecha_hasta})


@reportes_bp.route('/tareas/pdf')
@login_required
@requiere_permiso('reportes')
def tareas_pdf():
    """Generar PDF profesional de tareas"""
    estado = request.args.get('estado')
    usuario_id = request.args.get('usuario_id')

    # Obtener tareas
    if current_user.es_admin():
        tareas = TareaDAO.listar(estado=estado, usuario_id=int(usuario_id) if usuario_id else None)
    else:
        tareas = TareaDAO.listar(estado=estado, usuario_id=current_user.id)

    # Estadísticas
    total = len(tareas)
    completadas = len([t for t in tareas if t.estado == 'completada'])
    pendientes = len([t for t in tareas if t.estado == 'pendiente'])
    en_proceso = len([t for t in tareas if t.estado == 'en_proceso'])
    postergadas = len([t for t in tareas if t.estado == 'postergada'])

    # Crear PDF con header/footer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=1 * inch,
        bottomMargin=0.8 * inch,
        leftMargin=40,
        rightMargin=40
    )

    elements = []
    styles = getSampleStyleSheet()

    # Título principal
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        alignment=TA_CENTER,
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph('Reporte de Tareas', title_style))

    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#6b7280'),
        alignment=TA_CENTER,
        spaceAfter=25
    )
    elements.append(Paragraph(f'Total: {total} tareas encontradas', subtitle_style))

    # Tarjetas de estadísticas (4 columnas)
    stats_data = [
        ['COMPLETADAS', 'PENDIENTES', 'EN PROCESO', 'POSTERGADAS'],
        [str(completadas), str(pendientes), str(en_proceso), str(postergadas)]
    ]

    stats_table = Table(stats_data, colWidths=[2 * inch, 2 * inch, 2 * inch, 2 * inch])
    stats_table.setStyle(TableStyle([
        # Encabezados
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#10b981')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#f59e0b')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#06b6d4')),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#ef4444')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        # Valores
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f9fafb')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 18),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1f2937')),
        ('TOPPADDING', (0, 1), (-1, 1), 12),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))

    elements.append(stats_table)
    elements.append(Spacer(1, 0.4 * inch))

    # Tabla principal de tareas
    data = [['#', 'TÍTULO', 'ASIGNADO', 'ENTREGA', 'PRIORIDAD', 'ESTADO']]

    for t in tareas:
        # Determinar color de prioridad
        if t.prioridad == 'alta':
            prioridad_text = 'Alta'
        elif t.prioridad == 'media':
            prioridad_text = 'Media'
        else:
            prioridad_text = 'Baja'

        # Determinar estado
        if t.estado == 'completada':
            estado_text = 'Completada'
        elif t.estado == 'postergada':
            estado_text = 'Postergada'
        elif t.estado == 'en_proceso':
            estado_text = 'En proceso'
        else:
            estado_text = 'Pendiente'

        data.append([
            f'#{t.id}',
            t.titulo[:35],
            (t.usuario_nombre or 'Sin asignar')[:20],
            t.fecha_entrega.strftime('%d/%m/%Y') if t.fecha_entrega else '—',
            prioridad_text,
            estado_text
        ])

    table = Table(data, colWidths=[0.5 * inch, 2.8 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1.2 * inch])
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

        # Filas de datos
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID centrado
        ('ALIGN', (1, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Bordes y colores alternados
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
    ]))

    elements.append(table)

    # Construir PDF con header/footer
    doc.build(elements, onFirstPage=crear_header_footer, onLaterPages=crear_header_footer)

    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=f'tareas_{datetime.now().strftime("%Y%m%d")}.pdf')


@reportes_bp.route('/tareas/excel')
@login_required
@requiere_permiso('reportes')
def tareas_excel():
    """Generar Excel profesional de tareas"""
    estado = request.args.get('estado')
    usuario_id = request.args.get('usuario_id')

    # Obtener tareas
    if current_user.es_admin():
        tareas = TareaDAO.listar(estado=estado, usuario_id=int(usuario_id) if usuario_id else None)
    else:
        tareas = TareaDAO.listar(estado=estado, usuario_id=current_user.id)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tareas'

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'REPORTE DE TAREAS'
    ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Fecha
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(size=10, color='6B7280')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    # Encabezados de tabla
    headers_row = 4
    headers = ['ID', 'Título', 'Descripción', 'Asignado a', 'Fecha Entrega', 'Prioridad', 'Estado', 'Campaña']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=headers_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[headers_row].height = 25

    # Datos
    for idx, t in enumerate(tareas, start=headers_row + 1):
        row_data = [
            t.id,
            t.titulo,
            t.descripcion or '',
            t.usuario_nombre or 'Sin asignar',
            t.fecha_entrega.strftime('%d/%m/%Y') if t.fecha_entrega else '',
            t.prioridad.capitalize(),
            t.estado.replace('_', ' ').capitalize(),
            t.campana_nombre or ''
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Color alternado
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    # Ajustar anchos
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20

    # Guardar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'tareas_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ==================== REPORTES DE CAMPAÑAS ====================

@reportes_bp.route('/campanas')
@login_required
@requiere_permiso('reportes')
def campanas():
    """Vista previa de reporte de campañas"""
    campanas = CampanaDAO.listar()

    total = len(campanas)
    activas = len([c for c in campanas if c.status == 'activa'])
    finalizadas = len([c for c in campanas if c.status == 'finalizada'])
    presupuesto_total = sum([c.presupuesto for c in campanas if c.presupuesto])
    gastado_total = sum([c.gastado for c in campanas if c.gastado])

    return render_template('reportes/campanas.html',
                           campanas=campanas,
                           total=total,
                           activas=activas,
                           finalizadas=finalizadas,
                           presupuesto_total=presupuesto_total,
                           gastado_total=gastado_total)


@reportes_bp.route('/campanas/pdf')
@login_required
@requiere_permiso('reportes')
def campanas_pdf():
    """Generar PDF de campañas con estadísticas"""
    campanas = CampanaDAO.listar()

    total = len(campanas)
    activas = len([c for c in campanas if c.status == 'activa'])
    finalizadas = len([c for c in campanas if c.status == 'finalizada'])
    presupuesto_total = sum([c.presupuesto for c in campanas if c.presupuesto])
    gastado_total = sum([c.gastado for c in campanas if c.gastado])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1 * inch, bottomMargin=0.8 * inch, leftMargin=40,
                            rightMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                                 textColor=colors.HexColor('#1e40af'), alignment=TA_CENTER, spaceAfter=10,
                                 fontName='Helvetica-Bold')
    elements.append(Paragraph('Reporte de Campañas', title_style))

    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=11,
                                    textColor=colors.HexColor('#6b7280'), alignment=TA_CENTER, spaceAfter=25)
    elements.append(Paragraph(f'Total: {total} campañas', subtitle_style))

    # Tarjetas de estadísticas (5 columnas)
    stats_data = [
        ['TOTAL', 'ACTIVAS', 'FINALIZADAS', 'PRESUPUESTO', 'GASTADO'],
        [str(total), str(activas), str(finalizadas), f'${presupuesto_total:,.0f}', f'${gastado_total:,.0f}']
    ]

    stats_table = Table(stats_data, colWidths=[1.6 * inch] * 5)
    stats_table.setStyle(TableStyle([
        # Encabezados
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#3b82f6')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#10b981')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#64748b')),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#8b5cf6')),
        ('BACKGROUND', (4, 0), (4, 0), colors.HexColor('#ef4444')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        # Valores
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f9fafb')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 16),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1f2937')),
        ('TOPPADDING', (0, 1), (-1, 1), 12),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))

    elements.append(stats_table)
    elements.append(Spacer(1, 0.4 * inch))

    # Tabla principal
    data = [['#', 'NOMBRE', 'INICIO', 'FIN', 'PRESUPUESTO', 'GASTADO', 'ESTADO']]
    for c in campanas:
        data.append([
            f'#{c.id}',
            c.nombre[:30],
            c.fecha_inicio.strftime('%d/%m/%Y') if c.fecha_inicio else '—',
            c.fecha_fin.strftime('%d/%m/%Y') if c.fecha_fin else '—',
            f'${c.presupuesto:,.0f}' if c.presupuesto else '—',
            f'${c.gastado:,.0f}' if c.gastado else '$0',
            c.status.capitalize()
        ])

    table = Table(data, colWidths=[0.5 * inch, 2.5 * inch, 1 * inch, 1 * inch, 1.2 * inch, 1.2 * inch, 0.8 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
    ]))

    elements.append(table)
    doc.build(elements, onFirstPage=crear_header_footer, onLaterPages=crear_header_footer)

    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=f'campanas_{datetime.now().strftime("%Y%m%d")}.pdf')


@reportes_bp.route('/campanas/excel')
@login_required
@requiere_permiso('reportes')
def campanas_excel():
    """Generar Excel de campañas"""
    campanas = CampanaDAO.listar()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Campañas'

    ws.merge_cells('A1:I1')
    ws['A1'] = 'REPORTE DE CAMPAÑAS'
    ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers_row = 3
    headers = ['ID', 'Nombre', 'Descripción', 'Fecha Inicio', 'Fecha Fin', 'Presupuesto', 'Gastado', '% Ejecutado',
               'Estado']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=headers_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, c in enumerate(campanas, start=headers_row + 1):
        pct_gastado = ((c.gastado / c.presupuesto * 100) if c.presupuesto and c.gastado else 0)
        row_data = [
            c.id,
            c.nombre,
            c.descripcion or '',
            c.fecha_inicio.strftime('%d/%m/%Y') if c.fecha_inicio else '',
            c.fecha_fin.strftime('%d/%m/%Y') if c.fecha_fin else '',
            c.presupuesto or 0,
            c.gastado or 0,
            f'{pct_gastado:.1f}%',
            c.status
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='left' if col > 1 else 'center', vertical='center')

            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'campanas_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ==================== REPORTES DE INFLUENCERS ====================

@reportes_bp.route('/influencers')
@login_required
@requiere_permiso('reportes')
def influencers():
    """Vista previa de reporte de influencers"""
    estado = request.args.get('estado')
    influencers = InfluencerDAO.listar(estado=estado)

    total = len(influencers)
    nuevos = len([i for i in influencers if i.estado == 'nuevo'])
    trabajaron = len([i for i in influencers if i.estado == 'trabajaron'])
    sugeridos = len([i for i in influencers if i.estado == 'sugerido'])

    return render_template('reportes/influencers.html',
                           influencers=influencers,
                           total=total,
                           nuevos=nuevos,
                           trabajaron=trabajaron,
                           sugeridos=sugeridos,
                           filtro_estado=estado)


@reportes_bp.route('/influencers/pdf')
@login_required
@requiere_permiso('reportes')
def influencers_pdf():
    """Generar PDF de influencers con estadísticas"""
    estado = request.args.get('estado')
    influencers = InfluencerDAO.listar(estado=estado)

    total = len(influencers)
    nuevos = len([i for i in influencers if i.estado == 'nuevo'])
    trabajaron = len([i for i in influencers if i.estado == 'trabajaron'])
    sugeridos = len([i for i in influencers if i.estado == 'sugerido'])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1 * inch, bottomMargin=0.8 * inch, leftMargin=40,
                            rightMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                                 textColor=colors.HexColor('#1e40af'), alignment=TA_CENTER, spaceAfter=10,
                                 fontName='Helvetica-Bold')
    elements.append(Paragraph('Reporte de Influencers', title_style))

    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=11,
                                    textColor=colors.HexColor('#6b7280'), alignment=TA_CENTER, spaceAfter=25)
    elements.append(Paragraph(f'Total: {total} influencers', subtitle_style))

    # Tarjetas de estadísticas (4 columnas)
    stats_data = [
        ['TOTAL', 'NUEVOS', 'YA TRABAJARON', 'SUGERIDOS'],
        [str(total), str(nuevos), str(trabajaron), str(sugeridos)]
    ]

    stats_table = Table(stats_data, colWidths=[2 * inch, 2 * inch, 2 * inch, 2 * inch])
    stats_table.setStyle(TableStyle([
        # Encabezados
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#3b82f6')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#10b981')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#8b5cf6')),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#f59e0b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        # Valores
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f9fafb')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 18),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1f2937')),
        ('TOPPADDING', (0, 1), (-1, 1), 12),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))

    elements.append(stats_table)
    elements.append(Spacer(1, 0.4 * inch))

    # Tabla principal
    data = [['#', 'NOMBRE', 'HANDLE', 'WHATSAPP', 'ESTADO']]
    for i in influencers:
        data.append([
            f'#{i.id}',
            i.nombre[:30],
            (i.handle or '—')[:20],
            (i.whatsapp or '—')[:20],
            i.estado.capitalize()
        ])

    table = Table(data, colWidths=[0.5 * inch, 2.5 * inch, 2 * inch, 1.5 * inch, 1.5 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
    ]))

    elements.append(table)
    doc.build(elements, onFirstPage=crear_header_footer, onLaterPages=crear_header_footer)

    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=f'influencers_{datetime.now().strftime("%Y%m%d")}.pdf')


@reportes_bp.route('/influencers/excel')
@login_required
@requiere_permiso('reportes')
def influencers_excel():
    """Generar Excel de influencers"""
    estado = request.args.get('estado')
    influencers = InfluencerDAO.listar(estado=estado)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Influencers'

    ws.merge_cells('A1:G1')
    ws['A1'] = 'REPORTE DE INFLUENCERS'
    ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers_row = 3
    headers = ['ID', 'Nombre', 'Handle', 'Instagram', 'WhatsApp', 'Estado', 'Notas']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=headers_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, i in enumerate(influencers, start=headers_row + 1):
        row_data = [
            i.id,
            i.nombre,
            i.handle or '',
            i.url_ig or '',
            i.whatsapp or '',
            i.estado,
            i.notas or ''
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='left' if col > 1 else 'center', vertical='center')

            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'influencers_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ==================== REPORTES DE ECOMMERCE ====================

@reportes_bp.route('/ecommerce')
@login_required
@requiere_permiso('reportes')
def ecommerce():
    """Vista previa de reporte de ecommerce"""
    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)
    meta_mensual = request.args.get('meta', type=float, default=50000000)

    if not mes or not anio:
        hoy = date.today()
        mes = hoy.month
        anio = hoy.year

    # Obtener todas las ventas
    ventas_raw = EcommerceDAO.listar_ventas()

    # Filtrar por mes y año - maneja tanto objetos como diccionarios
    ventas = []
    dias_con_ventas = set()

    for v in ventas_raw:
        if isinstance(v, dict):
            fecha_venta = v.get('fecha')
        else:
            fecha_venta = v.fecha

        fecha_venta = _coerce_fecha(fecha_venta)
        if fecha_venta and fecha_venta.month == mes and fecha_venta.year == anio:
            ventas.append(v)
            dias_con_ventas.add(fecha_venta.day)

    # Calcular totales - CORREGIDO: convertir a float
    total_ventas = 0.0
    for v in ventas:
        if isinstance(v, dict):
            total_ventas += float(v.get('monto', 0))
        else:
            total_ventas += float(v.monto)

    cantidad_ventas = len(ventas)
    ticket_promedio = total_ventas / cantidad_ventas if cantidad_ventas > 0 else 0

    # Nuevas métricas
    faltante_meta = meta_mensual - total_ventas
    porcentaje_cumplimiento = (total_ventas / meta_mensual * 100) if meta_mensual > 0 else 0
    dias_activos = len(dias_con_ventas)
    promedio_diario = total_ventas / dias_activos if dias_activos > 0 else 0

    return render_template('reportes/ecommerce.html',
                           ventas=ventas,
                           mes=mes,
                           anio=anio,
                           meta_mensual=meta_mensual,
                           total_ventas=total_ventas,
                           cantidad_ventas=cantidad_ventas,
                           ticket_promedio=ticket_promedio,
                           faltante_meta=faltante_meta,
                           porcentaje_cumplimiento=porcentaje_cumplimiento,
                           dias_activos=dias_activos,
                           promedio_diario=promedio_diario)


@reportes_bp.route('/ecommerce/pdf')
@login_required
@requiere_permiso('reportes')
def ecommerce_pdf():
    """Generar PDF de ecommerce con estadísticas completas"""
    mes = request.args.get('mes', type=int, default=date.today().month)
    anio = request.args.get('anio', type=int, default=date.today().year)
    meta_mensual = request.args.get('meta', type=float, default=50000000)

    # Obtener todas las ventas y filtrar
    ventas_raw = EcommerceDAO.listar_ventas()
    ventas = []
    dias_con_ventas = set()

    for v in ventas_raw:
        if isinstance(v, dict):
            fecha_venta = v.get('fecha')
        else:
            fecha_venta = v.fecha

        fecha_venta = _coerce_fecha(fecha_venta)
        if fecha_venta and fecha_venta.month == mes and fecha_venta.year == anio:
            ventas.append(v)
            dias_con_ventas.add(fecha_venta.day)

    # Calcular estadísticas - CORREGIDO: convertir a float
    total_ventas = 0.0
    for v in ventas:
        if isinstance(v, dict):
            total_ventas += float(v.get('monto', 0))
        else:
            total_ventas += float(v.monto)

    cantidad_ventas = len(ventas)
    ticket_promedio = total_ventas / cantidad_ventas if cantidad_ventas > 0 else 0
    faltante_meta = meta_mensual - total_ventas
    porcentaje_cumplimiento = (total_ventas / meta_mensual * 100) if meta_mensual > 0 else 0
    dias_activos = len(dias_con_ventas)
    promedio_diario = total_ventas / dias_activos if dias_activos > 0 else 0

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1 * inch, bottomMargin=0.8 * inch, leftMargin=40,
                            rightMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre',
             'Noviembre', 'Diciembre']

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                                 textColor=colors.HexColor('#1e40af'), alignment=TA_CENTER, spaceAfter=10,
                                 fontName='Helvetica-Bold')
    elements.append(Paragraph(f'Reporte de Ecommerce - {meses[mes]} {anio}', title_style))

    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=11,
                                    textColor=colors.HexColor('#6b7280'), alignment=TA_CENTER, spaceAfter=25)
    elements.append(Paragraph(
        f'Total: {cantidad_ventas} ventas | Meta: ${meta_mensual:,.0f} | Cumplimiento: {porcentaje_cumplimiento:.1f}%',
        subtitle_style))

    # Tarjetas de estadísticas (6 métricas)
    stats_data = [
        ['META DEL MES', 'TOTAL VENDIDO',
         f'{"FALTA" if faltante_meta > 0 else "EXCEDENTE"}', '% CUMPLIMIENTO', 'DÍAS ACTIVOS', 'PROMEDIO DIARIO'],
        [f'${meta_mensual:,.0f}', f'${total_ventas:,.0f}', f'${abs(faltante_meta):,.0f}',
         f'{porcentaje_cumplimiento:.1f}%', str(dias_activos), f'${promedio_diario:,.0f}']
    ]

    stats_table = Table(stats_data, colWidths=[1.3 * inch] * 6)
    stats_table.setStyle(TableStyle([
        # Encabezados
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#8b5cf6')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#10b981')),
        ('BACKGROUND', (2, 0), (2, 0),
         colors.HexColor('#ef4444') if faltante_meta > 0 else colors.HexColor('#10b981')),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#3b82f6')),
        ('BACKGROUND', (4, 0), (4, 0), colors.HexColor('#8b5cf6')),
        ('BACKGROUND', (5, 0), (5, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        # Valores
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f9fafb')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1f2937')),
        ('TOPPADDING', (0, 1), (-1, 1), 12),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))

    elements.append(stats_table)
    elements.append(Spacer(1, 0.4 * inch))

    # Tabla principal
    data = [['#', 'FECHA', 'INFLUENCER', 'MONTO', 'NOTA']]
    for v in ventas:
        if isinstance(v, dict):
            v_id = v.get('id')
            v_fecha = v.get('fecha')
            v_influencer = v.get('influencer_nombre', 'Orgánico')
            v_monto = float(v.get('monto', 0))
            v_nota = v.get('nota', '')
        else:
            v_id = v.id
            v_fecha = v.fecha
            v_influencer = v.influencer_nombre or 'Orgánico'
            v_monto = float(v.monto)
            v_nota = v.nota or ''

        v_fecha = _coerce_fecha(v_fecha)
        data.append([
            f'#{v_id}',
            v_fecha.strftime('%d/%m/%Y') if v_fecha else '—',
            str(v_influencer)[:25],
            f'${v_monto:,.0f}',
            str(v_nota)[:30]
        ])

    table = Table(data, colWidths=[0.5 * inch, 1 * inch, 2.5 * inch, 1.2 * inch, 2.8 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
    ]))

    elements.append(table)
    doc.build(elements, onFirstPage=crear_header_footer, onLaterPages=crear_header_footer)

    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=f'ecommerce_{mes}_{anio}.pdf')


@reportes_bp.route('/ecommerce/excel')
@login_required
@requiere_permiso('reportes')
def ecommerce_excel():
    """Generar Excel de ecommerce"""
    mes = request.args.get('mes', type=int, default=date.today().month)
    anio = request.args.get('anio', type=int, default=date.today().year)

    # Obtener todas las ventas y filtrar
    ventas_raw = EcommerceDAO.listar_ventas()
    ventas = []
    for v in ventas_raw:
        if isinstance(v, dict):
            fecha_venta = v.get('fecha')
        else:
            fecha_venta = v.fecha

        fecha_venta = _coerce_fecha(fecha_venta)
        if fecha_venta and fecha_venta.month == mes and fecha_venta.year == anio:
            ventas.append(v)

    wb = Workbook()
    ws = wb.active
    ws.title = f'{mes}-{anio}'

    ws.merge_cells('A1:E1')
    ws['A1'] = 'REPORTE DE ECOMMERCE'
    ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers_row = 3
    headers = ['ID', 'Fecha', 'Influencer', 'Monto', 'Nota']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=headers_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, v in enumerate(ventas, start=headers_row + 1):
        # Maneja tanto dict como objeto
        if isinstance(v, dict):
            v_id = v.get('id')
            v_fecha = v.get('fecha')
            v_influencer = v.get('influencer_nombre', 'Orgánico')
            v_monto = v.get('monto', 0)
            v_nota = v.get('nota', '')
        else:
            v_id = v.id
            v_fecha = v.fecha
            v_influencer = v.influencer_nombre or 'Orgánico'
            v_monto = v.monto
            v_nota = v.nota or ''

        v_fecha = _coerce_fecha(v_fecha)
        row_data = [
            v_id,
            v_fecha.strftime('%d/%m/%Y') if v_fecha else '',
            v_influencer,
            v_monto,
            v_nota
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='left' if col > 1 else 'center', vertical='center')

            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    # Calcular total
    total = 0
    for v in ventas:
        if isinstance(v, dict):
            total += v.get('monto', 0)
        else:
            total += v.monto

    total_row = len(ventas) + 4
    ws[f'C{total_row}'] = 'TOTAL:'
    ws[f'C{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'] = total
    ws[f'D{total_row}'].font = Font(bold=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'ecommerce_{mes}_{anio}.xlsx')
